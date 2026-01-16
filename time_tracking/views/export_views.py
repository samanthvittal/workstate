"""
Export views for time tracking data in CSV, PDF, and Excel formats.

Provides export functionality for time entries and analytics reports
in multiple formats.
"""
import logging
import csv
import io
from datetime import datetime, date
from decimal import Decimal
from django.contrib.auth.mixins import LoginRequiredMixin
from django.http import HttpResponse, FileResponse
from django.views import View

from time_tracking.services.analytics import AnalyticsService

logger = logging.getLogger(__name__)


class ExportCSVView(LoginRequiredMixin, View):
    """
    Export time entries to CSV format.

    GET /time/export/csv/
    Query params: start_date, end_date
    """

    def get(self, request):
        """Handle GET request for CSV export."""
        try:
            # Get date range
            start_date_str = request.GET.get('start_date')
            end_date_str = request.GET.get('end_date')

            if not start_date_str or not end_date_str:
                # Default to current month
                today = date.today()
                start_date = today.replace(day=1)
                if today.month == 12:
                    end_date = today.replace(day=31)
                else:
                    next_month = today.replace(month=today.month + 1, day=1)
                    end_date = next_month - datetime.timedelta(days=1)
            else:
                try:
                    start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
                    end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
                except ValueError:
                    return HttpResponse("Invalid date format", status=400)

            # Get export data
            analytics = AnalyticsService(request.user)
            export_data = analytics.get_csv_export_data(start_date, end_date)

            # Create CSV response
            response = HttpResponse(content_type='text/csv')
            response['Content-Disposition'] = f'attachment; filename="time_entries_{start_date}_{end_date}.csv"'

            # Write CSV
            writer = csv.writer(response)

            # Write header
            writer.writerow([
                'Date',
                'Task',
                'Project',
                'Duration (hours)',
                'Start Time',
                'End Time',
                'Description',
                'Tags',
                'Billable',
                'Rate',
                'Currency',
                'Revenue',
            ])

            # Write data rows
            for row in export_data:
                # Format duration as hours with 2 decimal places
                duration_hours = ''
                if row['duration']:
                    duration_hours = f"{row['duration'].total_seconds() / 3600:.2f}"

                writer.writerow([
                    row['date'],
                    row['task'],
                    row['project'],
                    duration_hours,
                    row['start_time'],
                    row['end_time'],
                    row['description'],
                    row['tags'],
                    row['billable'],
                    row['rate'],
                    row['currency'],
                    f"{row['revenue']:.2f}" if row['revenue'] else '',
                ])

            return response

        except Exception as e:
            logger.error(f"Error exporting CSV: {e}", exc_info=True)
            return HttpResponse("An error occurred during export", status=500)


class ExportPDFView(LoginRequiredMixin, View):
    """
    Export analytics report to PDF format.

    GET /time/export/pdf/
    Query params: start_date, end_date
    """

    def get(self, request):
        """Handle GET request for PDF export."""
        try:
            from reportlab.lib.pagesizes import letter, A4
            from reportlab.lib import colors
            from reportlab.lib.units import inch
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT

            # Get date range
            start_date_str = request.GET.get('start_date')
            end_date_str = request.GET.get('end_date')

            if not start_date_str or not end_date_str:
                today = date.today()
                start_date = today.replace(day=1)
                if today.month == 12:
                    end_date = today.replace(day=31)
                else:
                    next_month = today.replace(month=today.month + 1, day=1)
                    end_date = next_month - datetime.timedelta(days=1)
            else:
                try:
                    start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
                    end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
                except ValueError:
                    return HttpResponse("Invalid date format", status=400)

            # Get analytics data
            analytics = AnalyticsService(request.user)
            summary = analytics.get_summary_statistics(start_date, end_date)
            project_breakdown = analytics.get_project_breakdown(start_date, end_date)
            task_breakdown = analytics.get_task_breakdown(start_date, end_date, limit=10)

            # Create PDF buffer
            buffer = io.BytesIO()
            doc = SimpleDocTemplate(buffer, pagesize=letter)
            story = []

            # Styles
            styles = getSampleStyleSheet()
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=24,
                textColor=colors.HexColor('#1f2937'),
                spaceAfter=30,
                alignment=TA_CENTER,
            )
            heading_style = ParagraphStyle(
                'CustomHeading',
                parent=styles['Heading2'],
                fontSize=16,
                textColor=colors.HexColor('#374151'),
                spaceAfter=12,
            )

            # Title
            title = Paragraph(f"Time Tracking Report", title_style)
            story.append(title)

            # Date range
            date_range = Paragraph(
                f"<b>Period:</b> {start_date.strftime('%B %d, %Y')} to {end_date.strftime('%B %d, %Y')}",
                styles['Normal']
            )
            story.append(date_range)
            story.append(Spacer(1, 20))

            # Summary statistics
            story.append(Paragraph("Summary Statistics", heading_style))

            summary_data = [
                ['Total Hours', f"{summary['total_hours'].total_seconds() / 3600:.2f}"],
                ['Billable Hours', f"{summary['billable_hours'].total_seconds() / 3600:.2f}"],
                ['Non-billable Hours', f"{summary['non_billable_hours'].total_seconds() / 3600:.2f}"],
                ['Total Revenue', f"${summary['total_revenue']:.2f}"],
                ['Total Entries', str(summary['total_entries'])],
            ]

            summary_table = Table(summary_data, colWidths=[3*inch, 2*inch])
            summary_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#f9fafb')),
                ('TEXTCOLOR', (0, 0), (-1, -1), colors.HexColor('#1f2937')),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 11),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
                ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#e5e7eb')),
            ]))
            story.append(summary_table)
            story.append(Spacer(1, 30))

            # Project breakdown
            if project_breakdown:
                story.append(Paragraph("Time by Project", heading_style))

                project_data = [['Project', 'Hours', 'Entries']]
                for item in project_breakdown[:10]:
                    project_data.append([
                        item['project_name'],
                        f"{item['total_duration'].total_seconds() / 3600:.2f}",
                        str(item['entry_count']),
                    ])

                project_table = Table(project_data, colWidths=[3*inch, 1.5*inch, 1.5*inch])
                project_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#3b82f6')),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                    ('ALIGN', (1, 0), (-1, -1), 'RIGHT'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, -1), 10),
                    ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
                    ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#f9fafb')),
                    ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#e5e7eb')),
                ]))
                story.append(project_table)
                story.append(Spacer(1, 30))

            # Task breakdown
            if task_breakdown:
                story.append(Paragraph("Top 10 Tasks by Time", heading_style))

                task_data = [['Task', 'Project', 'Hours']]
                for item in task_breakdown:
                    task_data.append([
                        item['task_name'][:40],
                        item['project_name'][:30],
                        f"{item['total_duration'].total_seconds() / 3600:.2f}",
                    ])

                task_table = Table(task_data, colWidths=[2.5*inch, 2*inch, 1.5*inch])
                task_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#3b82f6')),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                    ('ALIGN', (2, 0), (-1, -1), 'RIGHT'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, -1), 10),
                    ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
                    ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#f9fafb')),
                    ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#e5e7eb')),
                ]))
                story.append(task_table)

            # Build PDF
            doc.build(story)

            # Return PDF response
            buffer.seek(0)
            response = FileResponse(buffer, content_type='application/pdf')
            response['Content-Disposition'] = f'attachment; filename="time_report_{start_date}_{end_date}.pdf"'

            return response

        except ImportError:
            logger.error("reportlab not installed")
            return HttpResponse("PDF export requires reportlab package", status=500)
        except Exception as e:
            logger.error(f"Error exporting PDF: {e}", exc_info=True)
            return HttpResponse("An error occurred during PDF export", status=500)


class ExportExcelView(LoginRequiredMixin, View):
    """
    Export time entries to Excel format with multiple sheets.

    GET /time/export/excel/
    Query params: start_date, end_date
    """

    def get(self, request):
        """Handle GET request for Excel export."""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill
            from openpyxl.utils import get_column_letter

            # Get date range
            start_date_str = request.GET.get('start_date')
            end_date_str = request.GET.get('end_date')

            if not start_date_str or not end_date_str:
                today = date.today()
                start_date = today.replace(day=1)
                if today.month == 12:
                    end_date = today.replace(day=31)
                else:
                    next_month = today.replace(month=today.month + 1, day=1)
                    end_date = next_month - datetime.timedelta(days=1)
            else:
                try:
                    start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
                    end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
                except ValueError:
                    return HttpResponse("Invalid date format", status=400)

            # Get analytics data
            analytics = AnalyticsService(request.user)
            summary = analytics.get_summary_statistics(start_date, end_date)
            export_data = analytics.get_csv_export_data(start_date, end_date)
            project_breakdown = analytics.get_project_breakdown(start_date, end_date)
            task_breakdown = analytics.get_task_breakdown(start_date, end_date, limit=20)

            # Create workbook
            wb = Workbook()

            # Header styles
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="3B82F6", end_color="3B82F6", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")

            # Sheet 1: Summary
            ws_summary = wb.active
            ws_summary.title = "Summary"

            ws_summary['A1'] = "Time Tracking Report"
            ws_summary['A1'].font = Font(bold=True, size=16)
            ws_summary['A3'] = f"Period: {start_date} to {end_date}"

            ws_summary['A5'] = "Metric"
            ws_summary['B5'] = "Value"
            ws_summary['A5'].font = header_font
            ws_summary['B5'].font = header_font
            ws_summary['A5'].fill = header_fill
            ws_summary['B5'].fill = header_fill

            ws_summary['A6'] = "Total Hours"
            ws_summary['B6'] = f"{summary['total_hours'].total_seconds() / 3600:.2f}"
            ws_summary['A7'] = "Billable Hours"
            ws_summary['B7'] = f"{summary['billable_hours'].total_seconds() / 3600:.2f}"
            ws_summary['A8'] = "Non-billable Hours"
            ws_summary['B8'] = f"{summary['non_billable_hours'].total_seconds() / 3600:.2f}"
            ws_summary['A9'] = "Total Revenue"
            ws_summary['B9'] = f"${summary['total_revenue']:.2f}"
            ws_summary['A10'] = "Total Entries"
            ws_summary['B10'] = summary['total_entries']

            # Sheet 2: All Entries
            ws_entries = wb.create_sheet("All Entries")

            headers = ['Date', 'Task', 'Project', 'Duration (h)', 'Start Time', 'End Time',
                      'Description', 'Tags', 'Billable', 'Rate', 'Currency', 'Revenue']

            for col_num, header in enumerate(headers, 1):
                cell = ws_entries.cell(row=1, column=col_num)
                cell.value = header
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment

            for row_num, entry in enumerate(export_data, 2):
                ws_entries.cell(row=row_num, column=1).value = str(entry['date'])
                ws_entries.cell(row=row_num, column=2).value = entry['task']
                ws_entries.cell(row=row_num, column=3).value = entry['project']
                ws_entries.cell(row=row_num, column=4).value = f"{entry['duration'].total_seconds() / 3600:.2f}" if entry['duration'] else ''
                ws_entries.cell(row=row_num, column=5).value = entry['start_time']
                ws_entries.cell(row=row_num, column=6).value = entry['end_time']
                ws_entries.cell(row=row_num, column=7).value = entry['description']
                ws_entries.cell(row=row_num, column=8).value = entry['tags']
                ws_entries.cell(row=row_num, column=9).value = entry['billable']
                ws_entries.cell(row=row_num, column=10).value = str(entry['rate']) if entry['rate'] else ''
                ws_entries.cell(row=row_num, column=11).value = entry['currency']
                ws_entries.cell(row=row_num, column=12).value = f"{entry['revenue']:.2f}" if entry['revenue'] else ''

            # Sheet 3: By Project
            ws_projects = wb.create_sheet("By Project")

            project_headers = ['Project', 'Hours', 'Entries']
            for col_num, header in enumerate(project_headers, 1):
                cell = ws_projects.cell(row=1, column=col_num)
                cell.value = header
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment

            for row_num, item in enumerate(project_breakdown, 2):
                ws_projects.cell(row=row_num, column=1).value = item['project_name']
                ws_projects.cell(row=row_num, column=2).value = f"{item['total_duration'].total_seconds() / 3600:.2f}"
                ws_projects.cell(row=row_num, column=3).value = item['entry_count']

            # Sheet 4: By Task
            ws_tasks = wb.create_sheet("By Task")

            task_headers = ['Task', 'Project', 'Hours', 'Entries']
            for col_num, header in enumerate(task_headers, 1):
                cell = ws_tasks.cell(row=1, column=col_num)
                cell.value = header
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment

            for row_num, item in enumerate(task_breakdown, 2):
                ws_tasks.cell(row=row_num, column=1).value = item['task_name']
                ws_tasks.cell(row=row_num, column=2).value = item['project_name']
                ws_tasks.cell(row=row_num, column=3).value = f"{item['total_duration'].total_seconds() / 3600:.2f}"
                ws_tasks.cell(row=row_num, column=4).value = item['entry_count']

            # Adjust column widths
            for ws in [ws_summary, ws_entries, ws_projects, ws_tasks]:
                for column in ws.columns:
                    max_length = 0
                    column_letter = get_column_letter(column[0].column)
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(cell.value)
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    ws.column_dimensions[column_letter].width = adjusted_width

            # Save to buffer
            buffer = io.BytesIO()
            wb.save(buffer)
            buffer.seek(0)

            # Return Excel response
            response = HttpResponse(
                buffer.getvalue(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename="time_report_{start_date}_{end_date}.xlsx"'

            return response

        except ImportError:
            logger.error("openpyxl not installed")
            return HttpResponse("Excel export requires openpyxl package", status=500)
        except Exception as e:
            logger.error(f"Error exporting Excel: {e}", exc_info=True)
            return HttpResponse("An error occurred during Excel export", status=500)
